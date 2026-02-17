"""保全薪資計算結果匯出 Excel（與前端表格欄位一致）。"""
import io
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


# 表頭（與前端「計算結果」表格一致）
EXCEL_HEADERS = [
    "案場", "員工", "薪制", "總工時", "應發", "勞保", "健保", "團保", "自提6%", "扣款合計", "實發", "狀態",
    "領薪方式", "銀行代碼", "分行代碼", "銀行帳號",
]


def _pay_type_label(pt: Any) -> str:
    if pt == "monthly":
        return "月薪"
    if pt == "daily":
        return "日薪"
    if pt == "hourly":
        return "時薪"
    return str(pt) if pt else ""


def _write_headers(ws, row_idx: int) -> None:
    thin = Side(style="thin")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, h in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=row_idx, column=col, value=h)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def _write_data_row(ws, row_idx: int, row: Dict[str, Any]) -> None:
    salary_type = row.get("salary_type") or ""
    is_cash = salary_type == "領現"
    is_unset = salary_type == "未設定"
    bank_code = "—" if is_cash else ("" if is_unset else (row.get("bank_code") or ""))
    branch_code = "—" if is_cash else ("" if is_unset else (row.get("branch_code") or ""))
    account_number = "—" if is_cash else ("" if is_unset else (row.get("account_number") or ""))
    ws.cell(row=row_idx, column=1, value=row.get("site") or "")
    ws.cell(row=row_idx, column=2, value=row.get("employee") or "")
    ws.cell(row=row_idx, column=3, value=_pay_type_label(row.get("pay_type")))
    ws.cell(row=row_idx, column=4, value=row.get("total_hours"))
    ws.cell(row=row_idx, column=5, value=row.get("gross_salary") or row.get("total_salary"))
    ws.cell(row=row_idx, column=6, value=row.get("labor_insurance_employee"))
    ws.cell(row=row_idx, column=7, value=row.get("health_insurance_employee"))
    ws.cell(row=row_idx, column=8, value=row.get("group_insurance"))
    ws.cell(row=row_idx, column=9, value=row.get("self_pension_6"))
    ws.cell(row=row_idx, column=10, value=row.get("deductions_total"))
    ws.cell(row=row_idx, column=11, value=row.get("net_salary") or row.get("total_salary"))
    ws.cell(row=row_idx, column=12, value=row.get("status") or "")
    ws.cell(row=row_idx, column=13, value=salary_type)
    ws.cell(row=row_idx, column=14, value=bank_code)
    ws.cell(row=row_idx, column=15, value=branch_code)
    ws.cell(row=row_idx, column=16, value=account_number)


def _apply_default_width(ws) -> None:
    for col in range(1, len(EXCEL_HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14


def build_payroll_excel(results: List[Dict[str, Any]], sheet_name: str = "薪資計算結果") -> bytes:
    """
    依 results（與 API 回傳的單筆結構一致）產生 Excel 二進位內容。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel 表單名稱長度限制

    _write_headers(ws, 1)
    for row_idx, row in enumerate(results, start=2):
        _write_data_row(ws, row_idx, row)
    _apply_default_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_payroll_excel_grouped(results: List[Dict[str, Any]], stats: Dict[str, int]) -> bytes:
    """
    依領薪方式分類輸出多 Sheet（固定 7 個工作表）。
    第一列：分類名稱與人數；第二列：表頭（有資料）或「無資料」（0 筆）。
    """
    wb = Workbook()
    sheet_order = [
        "全部顯示",
        "領現",
        "保全一銀",
        "公寓一銀",
        "史密斯一銀",
        "其他銀行",
        "未設定",
    ]
    stat_count = {
        "領現": int(stats.get("cash", 0) or 0),
        "保全一銀": int(stats.get("sec_first", 0) or 0),
        "公寓一銀": int(stats.get("apt_first", 0) or 0),
        "史密斯一銀": int(stats.get("smith_first", 0) or 0),
        "其他銀行": int(stats.get("other_bank", 0) or 0),
        "未設定": int(stats.get("unset", 0) or 0),
    }
    groups = {
        "全部顯示": list(results),
        "領現": [r for r in results if (r.get("salary_type") or "未設定") == "領現"],
        "保全一銀": [r for r in results if (r.get("salary_type") or "未設定") == "保全一銀"],
        "公寓一銀": [r for r in results if (r.get("salary_type") or "未設定") == "公寓一銀"],
        "史密斯一銀": [r for r in results if (r.get("salary_type") or "未設定") == "史密斯一銀"],
        "其他銀行": [r for r in results if (r.get("salary_type") or "未設定") == "其他銀行"],
        "未設定": [r for r in results if (r.get("salary_type") or "未設定") == "未設定"],
    }

    for idx, sheet_name in enumerate(sheet_order):
        ws = wb.active if idx == 0 else wb.create_sheet(title=sheet_name)
        ws.title = sheet_name
        rows = groups[sheet_name]
        count = len(rows) if sheet_name == "全部顯示" else stat_count[sheet_name]
        ws.cell(row=1, column=1, value=f"{sheet_name}（{count}人）")
        _apply_default_width(ws)

        if not rows:
            ws.cell(row=2, column=1, value="無資料")
            continue

        _write_headers(ws, 2)
        for row_idx, row in enumerate(rows, start=3):
            _write_data_row(ws, row_idx, row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
