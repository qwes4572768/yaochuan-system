"""人事資料自動備份：產生 Excel、寫入 server/backup/hr、僅保留最近 N 份。含檔案鎖防多實例重複執行。"""
import re
import time
from datetime import datetime
from pathlib import Path
from io import BytesIO
from decimal import Decimal

from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook

from app.config import settings, BASE_DIR
from app import crud

# 白名單：僅允許本系統產出的備份檔名（防路徑穿越）
BACKUP_FILENAME_PATTERN = re.compile(r"^hr_backup_\d{8}_\d{6}\.xlsx$")
from app.sensitive import employee_to_read_dict, dependent_to_read_dict

EMPLOYEE_COLUMNS = [
    "id", "name", "birth_date", "national_id", "reg_address", "live_address",
    "live_same_as_reg", "salary_type", "salary_value", "insured_salary_level",
    "enroll_date", "cancel_date", "dependent_count", "safety_pdf_path",
    "contract_84_1_pdf_path", "notes", "created_at", "updated_at",
]
DEPENDENT_COLUMNS = [
    "id", "employee_id", "name", "birth_date", "national_id", "relation",
    "city", "is_disabled", "disability_level", "notes", "created_at", "updated_at",
]


def _cell_value(v) -> str:
    from datetime import date
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.isoformat() if hasattr(v, "isoformat") else str(v)
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, Decimal):
        return str(v)
    return str(v)


def build_hr_backup_buffer(employees: list) -> tuple[BytesIO, str]:
    """依員工列表產生完整人事 Excel，回傳 (BytesIO, 檔名)。"""
    rows_emp = []
    rows_dep = []
    for emp in employees:
        d = employee_to_read_dict(emp, reveal_sensitive=True)
        row = [d.get(k) for k in EMPLOYEE_COLUMNS]
        rows_emp.append(row)
        for dep in emp.dependents or []:
            dd = dependent_to_read_dict(dep, reveal_sensitive=True)
            rows_dep.append([dd.get(k) for k in DEPENDENT_COLUMNS])

    wb = Workbook()
    ws_emp = wb.active
    ws_emp.title = "employees"
    for col, name in enumerate(EMPLOYEE_COLUMNS, 1):
        ws_emp.cell(row=1, column=col, value=name)
    for r, row in enumerate(rows_emp, 2):
        for c, val in enumerate(row, 1):
            ws_emp.cell(row=r, column=c, value=_cell_value(val) if val is not None else "")
    ws_dep = wb.create_sheet("dependents")
    for col, name in enumerate(DEPENDENT_COLUMNS, 1):
        ws_dep.cell(row=1, column=col, value=name)
    for r, row in enumerate(rows_dep, 2):
        for c, val in enumerate(row, 1):
            ws_dep.cell(row=r, column=c, value=_cell_value(val) if val is not None else "")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"hr_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buf, filename


def _get_backup_dir() -> Path:
    """備份目錄：絕對路徑或以 backend 專案根 BASE_DIR 為基準。"""
    p = settings.backup_dir
    if not p.is_absolute():
        p = BASE_DIR / p
    return p


def _prune_old_backups(backup_dir: Path, keep: int) -> None:
    """僅保留最近 keep 份 hr_backup_*.xlsx，刪除其餘。"""
    files = sorted(
        backup_dir.glob("hr_backup_*.xlsx"),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    for f in files[keep:]:
        try:
            f.unlink()
        except OSError:
            pass


def _acquire_backup_lock(backup_dir: Path, timeout_seconds: int = 600) -> bool:
    """取得備份鎖（檔案鎖）。同一時間僅一實例可執行。鎖檔逾時視為 stale 可覆蓋。回傳是否取得鎖。"""
    lock_path = backup_dir / ".hr_backup.lock"
    now = time.time()
    try:
        lock_path.touch(exist_ok=False)
        return True
    except FileExistsError:
        try:
            if now - lock_path.stat().st_mtime > timeout_seconds:
                lock_path.unlink()
                lock_path.touch(exist_ok=False)
                return True
        except OSError:
            pass
        return False


def _release_backup_lock(backup_dir: Path) -> None:
    try:
        (backup_dir / ".hr_backup.lock").unlink(missing_ok=True)
    except OSError:
        pass


async def run_scheduled_backup() -> str | None:
    """執行一次自動備份：寫入 server/backup/hr，並清理僅保留最近 N 份。多 worker 時以檔案鎖確保只跑一實例。回傳寫入檔名或 None。"""
    from app.database import AsyncSessionLocal
    backup_dir = _get_backup_dir()
    backup_dir.mkdir(parents=True, exist_ok=True)

    if not _acquire_backup_lock(backup_dir):
        return None

    try:
        async with AsyncSessionLocal() as db:
            employees = await crud.list_employees(
                db, skip=0, limit=100000, load_dependents=True, search=None
            )
        buf, filename = build_hr_backup_buffer(employees)
        path = backup_dir / filename
        path.write_bytes(buf.getvalue())
        _prune_old_backups(backup_dir, settings.backup_retention_count)
        return filename
    finally:
        _release_backup_lock(backup_dir)


def list_backup_files() -> list[dict]:
    """列出 backup_dir 內所有 hr_backup_*.xlsx，依修改時間由新到舊。"""
    backup_dir = _get_backup_dir()
    if not backup_dir.exists():
        return []
    files = sorted(
        backup_dir.glob("hr_backup_*.xlsx"),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    return [
        {
            "filename": f.name,
            "created_at": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
            "size": f.stat().st_size,
        }
        for f in files
    ]


def get_backup_path(filename: str) -> Path | None:
    """白名單：僅 hr_backup_YYYYMMDD_HHMMSS.xlsx，防路徑穿越。回傳完整路徑或 None。"""
    if not filename or not BACKUP_FILENAME_PATTERN.match(filename):
        return None
    if ".." in filename or "/" in filename or "\\" in filename:
        return None
    path = _get_backup_dir() / filename
    return path if path.is_file() else None
