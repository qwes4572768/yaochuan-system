"""解析勞健保級距表 Excel：格式為「雇主」與「員工」兩塊表格左右並排（第一個 sheet）。

欄位對應（0-based）：
  A=0 序號/文字（部分工時、標題、空白）
  B=1 級距金額（insured_salary_level）主 key
  C=2 勞保(雇主), D=3 健保(雇主), E=4 職災(雇主), F=5 勞退6%(雇主)
  G=6 雇主合計（不採用）
  H=7 級距金額(員工側)
  I=8 勞保(員工), J=9 健保(員工)
  K=10 員工合計（不採用）
團保不從 Excel 匯入，固定 0；由系統設定/公式處理。
"""
from io import BytesIO
import re
import logging
from typing import List, Dict, Any, Optional, Tuple

from decimal import Decimal, InvalidOperation

logger = logging.getLogger(__name__)

# 級距欄：先判斷是否為數字再轉，絕不直接 int(Decimal(str(level_raw)))
LEVEL_NUMERIC_PATTERN = re.compile(r"^-?\d+\.?\d*$")

# 固定欄位索引（0-based）：B=1, C=2, D=3, E=4, F=5, I=8, J=9
COL_LEVEL = 1
COL_LABOR_EMPLOYER = 2
COL_HEALTH_EMPLOYER = 3
COL_OCC = 4
COL_PENSION = 5
COL_LABOR_EMPLOYEE = 8
COL_HEALTH_EMPLOYEE = 9


def _normalize_number_str(raw: Any) -> str:
    """將原始值轉成字串並移除逗號、空白、'元' 等。"""
    if raw is None:
        return ""
    s = str(raw).strip()
    for remove in (",", " ", "　", "元", "NT$", "NTD", "N.T.", "$", "¥", "－", "—"):
        s = s.replace(remove, "")
    return s.strip()


def _is_level_number(raw: Any) -> bool:
    """判斷 B 欄是否為可解析的級距數字（不進行轉換，只判斷）。"""
    if raw is None:
        return False
    if isinstance(raw, (int, float)):
        try:
            if isinstance(raw, float) and (raw != raw or raw < 0):
                return False
            if isinstance(raw, int) and raw < 0:
                return False
            return True
        except (TypeError, ValueError):
            return False
    s = _normalize_number_str(raw)
    if not s or s in ("-", "—", "－"):
        return False
    if s.lower() in ("nan", "na", "部分工時", "級距", "級距金額"):
        return False
    s_clean = re.sub(r"[^\d.\-]", "", s)
    if not s_clean:
        return False
    return bool(LEVEL_NUMERIC_PATTERN.match(s_clean) or re.match(r"^\d+\.?\d*$", s_clean))


def _parse_level_safe(raw: Any) -> Optional[int]:
    """
    解析級距為整數。僅在已確認為數字格式時呼叫；支援 42000、42000.0、'42,000'。
    若轉換失敗回傳 None（呼叫端應先以 _is_level_number 過濾，此處僅防萬一）。
    """
    if raw is None:
        return None
    if isinstance(raw, int):
        return raw if raw >= 0 else None
    if isinstance(raw, float):
        try:
            if raw != raw or raw < 0:
                return None
            return int(round(raw))
        except (ValueError, TypeError):
            return None
    s = _normalize_number_str(raw)
    s_clean = re.sub(r"[^\d.\-]", "", s)
    if not s_clean:
        return None
    try:
        f = float(s_clean)
        if f < 0:
            return None
        return int(round(f))
    except (ValueError, TypeError, InvalidOperation):
        return None


def _parse_decimal_safe(raw: Any) -> Decimal:
    """解析數字欄位，空白/非數字視為 0；不拋錯。"""
    if raw is None:
        return Decimal("0")
    s = _normalize_number_str(raw)
    if not s or s in ("-", "—", "－"):
        return Decimal("0")
    s_clean = re.sub(r"[^\d.\-]", "", s)
    if not s_clean:
        return Decimal("0")
    try:
        return Decimal(s_clean)
    except (ValueError, InvalidOperation, TypeError):
        return Decimal("0")


def parse_bracket_excel(content: bytes) -> Dict[str, Any]:
    """
    解析「雇主／員工左右並排」格式之級距表 Excel。
    從第 1 個 sheet 讀取，逐列取 B 欄為級距；若 B 不是數字則跳過該列。
    回傳：rows, errors（最多 20 筆）, header_row_index（此格式不找表頭，固定為 None 或 0）。
    """
    from openpyxl import load_workbook

    result: Dict[str, Any] = {"rows": [], "errors": [], "header_row_index": 0}
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        result["errors"].append({"message": "Excel 無有效工作表"})
        return result
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        result["errors"].append({"message": "Excel 無資料列"})
        return result

    for row_idx, row in enumerate(rows):
        excel_row_1based = row_idx + 1
        if row is None or len(row) <= COL_LEVEL:
            continue
        # B 欄 = 級距
        level_raw = row[COL_LEVEL] if COL_LEVEL < len(row) else None
        if excel_row_1based <= 10:
            logger.debug("bracket_excel row %s B(level_raw)=%r", excel_row_1based, level_raw)

        if not _is_level_number(level_raw):
            continue
        level = _parse_level_safe(level_raw)
        if level is None:
            continue

        def cell(i: int) -> Any:
            return row[i] if i < len(row) else None

        labor_employer = _parse_decimal_safe(cell(COL_LABOR_EMPLOYER))
        health_employer = _parse_decimal_safe(cell(COL_HEALTH_EMPLOYER))
        occ = _parse_decimal_safe(cell(COL_OCC))
        labor_pension = _parse_decimal_safe(cell(COL_PENSION))
        labor_employee = _parse_decimal_safe(cell(COL_LABOR_EMPLOYEE))
        health_employee = _parse_decimal_safe(cell(COL_HEALTH_EMPLOYEE))
        group_insurance = Decimal("0")

        result["rows"].append({
            "insured_salary_level": level,
            "labor_employer": labor_employer,
            "labor_employee": labor_employee,
            "health_employer": health_employer,
            "health_employee": health_employee,
            "occupational_accident": occ,
            "labor_pension": labor_pension,
            "group_insurance": group_insurance,
        })

    return result
