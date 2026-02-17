"""勞健保/職災/團保/勞退試算 API（級距表為依據）；級距下拉 API；保險結果落表供會計；Excel 試算檔上傳"""
from io import BytesIO
from decimal import Decimal
from typing import List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app import crud
from app.crud import INSURANCE_ITEM_TYPES, delete_insurance_monthly_results_for_month, upsert_insurance_monthly_result
from app.schemas import (
    InsuranceEstimateRequest,
    InsuranceEstimateResponse,
    ItemBreakdown,
    SalaryBracketItem,
    InsuranceMonthlyResultRead,
)
from app.services.insurance_calc import get_brackets, salary_to_level
from app.config import settings

router = APIRouter(prefix="/api/insurance", tags=["insurance"])


def _brackets_from_db(imp) -> list[SalaryBracketItem]:
    """由最新匯入的 insurance_brackets 組出 [ { level, low, high } ]：level 排序後，low=前一級+1（首筆 1），high=本級距（末筆 999999）。"""
    if not imp or not getattr(imp, "brackets", None) or not imp.brackets:
        return []
    levels = sorted([int(b.insured_salary_level) for b in imp.brackets])
    out = []
    for i, lev in enumerate(levels):
        low = 1 if i == 0 else levels[i - 1] + 1
        high = lev if i < len(levels) - 1 else 999999
        out.append(SalaryBracketItem(level=Decimal(lev), low=low, high=high))
    return out


@router.get("/brackets", response_model=list[SalaryBracketItem])
async def list_brackets(db: AsyncSession = Depends(get_db)):
    """取得投保薪資級距列表：以 DB 最新匯入的 insurance_brackets 為準；無匯入時 fallback 至 YAML。"""
    imp = await crud.get_latest_bracket_import(db)
    from_db = _brackets_from_db(imp)
    if from_db:
        return from_db
    rules = await crud.get_all_insurance_rules(db)
    raw = get_brackets(rules)
    return [SalaryBracketItem(level=Decimal(str(b[2])), low=int(b[0]), high=int(b[1])) for b in raw]


@router.get("/salary-to-level")
async def map_salary_to_level(
    salary: int,
    db: AsyncSession = Depends(get_db),
):
    """輸入金額後自動對應級距：與下拉選單同源（DB 最新匯入）；無匯入時用 YAML。"""
    imp = await crud.get_latest_bracket_import(db)
    items = _brackets_from_db(imp)
    if items:
        for it in items:
            if it.low <= salary <= it.high:
                return {"salary": salary, "insured_salary_level": float(it.level)}
        if salary > items[-1].high:
            return {"salary": salary, "insured_salary_level": float(items[-1].level)}
        return {"salary": salary, "insured_salary_level": float(items[0].level)}
    rules = await crud.get_all_insurance_rules(db)
    level = salary_to_level(Decimal(str(salary)), rules)
    return {"salary": salary, "insured_salary_level": float(level)}


def _build_persons(emp) -> list:
    """由員工 + 眷屬組出健保減免用 persons（姓名、本人/眷屬、出生日、縣市、身障等級）"""
    persons = []
    persons.append({
        "name": emp.name,
        "is_employee": True,
        "birth_date": emp.birth_date.isoformat() if emp.birth_date else None,
        "city": None,
        "disability_level": None,
    })
    for d in emp.dependents or []:
        persons.append({
            "name": d.name,
            "is_employee": False,
            "birth_date": d.birth_date.isoformat() if d.birth_date else None,
            "city": d.city,
            "disability_level": d.disability_level if d.is_disabled else None,
        })
    return persons


@router.post("/estimate", response_model=InsuranceEstimateResponse)
async def insurance_estimate(
    body: InsuranceEstimateRequest,
    db: AsyncSession = Depends(get_db),
):
    """
    試算：以「級距表匯入」為唯一依據。有匯入級距表時依投保級距查表回傳公司/員工/合計；
    未匯入或找不到級距時回傳明確錯誤。可傳 employee_id 自動帶入投保級距。
    """
    level = body.insured_salary_level
    dep_count = body.dependent_count

    emp = None
    if body.employee_id:
        emp = await crud.get_employee(db, body.employee_id)
        if emp and level is None:
            level = emp.insured_salary_level
        if emp and dep_count == 0:
            dep_count = emp.dependent_count if emp.dependent_count is not None else len(emp.dependents or [])

    level_int = None
    if level is not None:
        try:
            level_int = int(Decimal(str(level)))
        except (ValueError, TypeError):
            pass

    if level_int is None:
        raise HTTPException(status_code=400, detail="請提供投保級距或選擇已設定投保級距的員工")

    imp = await crud.get_latest_bracket_import(db)
    if not imp or not getattr(imp, "brackets", None):
        raise HTTPException(status_code=400, detail="查無級距，請先匯入級距表或確認員工級距金額")
    bracket = await crud.get_bracket_by_level(db, imp.id, level_int)
    if not bracket:
        raise HTTPException(status_code=400, detail="查無級距，請先匯入級距表或確認員工級距金額")

    labor_employer = bracket.labor_employer
    labor_employee = bracket.labor_employee
    health_employer = bracket.health_employer
    health_employee = bracket.health_employee
    occ = bracket.occupational_accident
    pension = bracket.labor_pension
    # 團保：固定月費（config），不按天數比例；全由員工負擔（不從 Excel 級距表讀取）
    group_monthly = Decimal(str(settings.group_insurance_monthly_fee))
    group_employer = Decimal("0")
    group_employee = group_monthly
    group_total = group_monthly
    total_employer = labor_employer + health_employer + occ + pension + group_employer
    total_employee = labor_employee + health_employee + group_employee
    total = total_employer + total_employee
    # 有傳 employee_id 時以 DB 儲存的 pension_self_6 為準，否則用 request body
    use_pension_self_6 = bool(getattr(emp, "pension_self_6", False)) if (body.employee_id and emp) else body.pension_self_6
    pension_self_6_item = None
    if use_pension_self_6:
        pension_self_6_item = ItemBreakdown(name="自提6%", employer=Decimal("0"), employee=pension, total=pension)
        total_employee = total_employee + pension
        total = total_employer + total_employee
    imported_at_str = imp.imported_at.strftime("%Y-%m-%d %H:%M") if imp.imported_at else ""

    return InsuranceEstimateResponse(
        insured_salary_level=Decimal(level_int),
        labor_insurance=ItemBreakdown(
            name="勞保", employer=labor_employer, employee=labor_employee, total=labor_employer + labor_employee
        ),
        health_insurance=ItemBreakdown(
            name="健保", employer=health_employer, employee=health_employee, total=health_employer + health_employee
        ),
        health_insurance_breakdown=None,
        occupational_accident=ItemBreakdown(name="職災", employer=occ, employee=Decimal("0"), total=occ),
        labor_pension=ItemBreakdown(name="勞退6%", employer=pension, employee=Decimal("0"), total=pension),
        group_insurance=ItemBreakdown(name="團保", employer=group_employer, employee=group_employee, total=group_total),
        pension_self_6=pension_self_6_item,
        total_employer=total_employer,
        total_employee=total_employee,
        total=total,
        dependent_count=dep_count or 0,
        from_bracket_table=True,
        bracket_source={"file_name": imp.file_name or "", "imported_at": imported_at_str},
    )


def _parse_excel_totals(content: bytes) -> Tuple[Decimal, Decimal, Decimal]:
    """
    從試算 Excel 解析「合計」列，回傳 (公司負擔, 員工負擔, 合計)。
    預期表頭含：項目/名稱 + 雇主/公司負擔 + 員工/員工負擔/個人負擔 + 小計/合計。
    """
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="Excel 無有效工作表")
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 無資料")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_item = None
    col_employer = None
    col_employee = None
    col_total = None
    for i, h in enumerate(header):
        if not h:
            continue
        if "項目" in h or "名稱" in h:
            col_item = i
        if "雇主" in h or "公司負擔" in h:
            col_employer = i
        if "員工" in h or "個人負擔" in h or "員工負擔" in h:
            col_employee = i
        if "小計" in h or "合計" in h:
            col_total = i
    if col_employer is None or col_employee is None or col_total is None:
        raise HTTPException(
            status_code=400,
            detail="Excel 需含「雇主/公司負擔」「員工/員工負擔」「小計/合計」欄位",
        )
    total_employer = Decimal("0")
    total_employee = Decimal("0")
    total = Decimal("0")
    for row in rows[1:]:
        if not row:
            continue
        first_cell = str(row[col_item] or "").strip() if col_item is not None and col_item < len(row) else ""
        if "合計" in first_cell:
            try:
                total_employer = Decimal(str(row[col_employer] or 0))
                total_employee = Decimal(str(row[col_employee] or 0))
                total = Decimal(str(row[col_total] or 0))
            except Exception:
                pass
            break
    if total == 0 and total_employer == 0 and total_employee == 0:
        raise HTTPException(status_code=400, detail="Excel 中未找到「合計」列或數值為空")
    return total_employer, total_employee, total


@router.post("/estimate-from-excel", response_model=InsuranceEstimateResponse)
async def insurance_estimate_from_excel(
    employee_id: int = Query(..., description="員工 ID"),
    year: int = Query(..., description="年度"),
    month: int = Query(..., ge=1, le=12, description="月份"),
    file: UploadFile = File(..., description="試算 Excel 檔（需含合計列：公司負擔、員工負擔、合計）"),
    db: AsyncSession = Depends(get_db),
):
    """
    上傳 Excel 試算檔，系統僅解析並回傳 Excel 中的「公司負擔、員工負擔、合計」，
    不再用系統內建費率計算。公司負擔以 Excel 為準（含勞退 6% 等所有項目）。
    """
    if not file.filename or not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="僅接受 .xlsx 或 .xls 試算檔")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="檔案不得超過 10MB")
    total_employer, total_employee, total = _parse_excel_totals(content)
    emp = await crud.get_employee(db, employee_id)
    level = Decimal("0")
    dep_count = 0
    if emp:
        level = emp.insured_salary_level or Decimal("0")
        dep_count = emp.dependent_count if emp.dependent_count is not None else len(emp.dependents or [])
    return InsuranceEstimateResponse(
        insured_salary_level=level,
        labor_insurance=ItemBreakdown(name="勞保", employer=Decimal("0"), employee=Decimal("0"), total=Decimal("0")),
        health_insurance=ItemBreakdown(name="健保", employer=Decimal("0"), employee=Decimal("0"), total=Decimal("0")),
        health_insurance_breakdown=None,
        occupational_accident=ItemBreakdown(name="職災", employer=Decimal("0"), employee=Decimal("0"), total=Decimal("0")),
        labor_pension=ItemBreakdown(name="勞退6%", employer=Decimal("0"), employee=Decimal("0"), total=Decimal("0")),
        group_insurance=ItemBreakdown(name="團保", employer=Decimal("0"), employee=Decimal("0"), total=Decimal("0")),
        pension_self_6=None,
        total_employer=total_employer,
        total_employee=total_employee,
        total=total,
        dependent_count=dep_count,
        from_excel=True,
    )


# ---------- 保險結果落表（會計抓 company cost） ----------
@router.post("/monthly-result/generate")
async def generate_monthly_insurance_result(
    year: int = Query(..., description="年度"),
    month: int = Query(..., ge=1, le=12, description="月份"),
    overwrite: bool = Query(True, description="是否覆蓋既有結果"),
    db: AsyncSession = Depends(get_db),
):
    """
    依指定年月、以級距表為依據計算所有員工保險費用並落表 insurance_monthly_result。
    會計可依 year_month 查詢 GET /api/insurance/monthly-result 取得 company cost。
    """
    year_month = year * 100 + month
    if overwrite:
        await delete_insurance_monthly_results_for_month(db, year_month)
    imp = await crud.get_latest_bracket_import(db)
    if not imp or not getattr(imp, "brackets", None):
        raise HTTPException(
            status_code=400,
            detail="級距表尚未匯入，請先至「級距表匯入」上傳 Excel 後再產生月結果",
        )
    employees = await crud.list_employees(db, skip=0, limit=10000, load_dependents=True)
    count = 0
    for e in employees:
        level_raw = e.insured_salary_level
        if level_raw is None or level_raw <= 0:
            continue
        try:
            level_int = int(Decimal(str(level_raw)))
        except (ValueError, TypeError):
            continue
        bracket = await crud.get_bracket_by_level(db, imp.id, level_int)
        if not bracket:
            continue
        lab_gov = None
        await upsert_insurance_monthly_result(
            db, e.id, year_month, "labor_insurance",
            bracket.labor_employee, bracket.labor_employer, lab_gov,
        )
        await upsert_insurance_monthly_result(
            db, e.id, year_month, "health_insurance",
            bracket.health_employee, bracket.health_employer, None,
        )
        await upsert_insurance_monthly_result(
            db, e.id, year_month, "occupational_accident",
            Decimal("0"), bracket.occupational_accident, None,
        )
        await upsert_insurance_monthly_result(
            db, e.id, year_month, "labor_pension",
            Decimal("0"), bracket.labor_pension, None,
        )
        # 團保固定月費由 config，全由員工負擔（不從級距表）
        _group_monthly = Decimal(str(settings.group_insurance_monthly_fee))
        await upsert_insurance_monthly_result(
            db, e.id, year_month, "group_insurance",
            _group_monthly, Decimal("0"), None,
        )
        count += 1
    await db.commit()
    return {"year": year, "month": month, "year_month": year_month, "employees_processed": count}


@router.get("/monthly-result", response_model=List[InsuranceMonthlyResultRead])
async def list_monthly_insurance_result(
    year_month: int = Query(..., description="西元年月，如 202501"),
    employee_id: Optional[int] = Query(None, description="篩選單一員工"),
    db: AsyncSession = Depends(get_db),
):
    """查詢保險結果落表（會計抓 company cost）。"""
    rows = await crud.list_insurance_monthly_results(db, year_month, employee_id=employee_id)
    return [InsuranceMonthlyResultRead.model_validate(r) for r in rows]