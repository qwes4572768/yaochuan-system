"""報表：員工清單、眷屬清單、當月公司負擔明細 - 匯出 Excel"""
from io import BytesIO
from datetime import date
from decimal import Decimal
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from app.database import get_db
from app import crud
from app.crypto import decrypt
from app.services.insurance_calc import estimate_insurance

router = APIRouter(prefix="/api/reports", tags=["reports"])


def _style_header(ws):
    thin = Side(style="thin")
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


@router.get("/export/employees")
async def export_employees_excel(db: AsyncSession = Depends(get_db)):
    """匯出員工清單 Excel"""
    employees = await crud.list_employees(db, skip=0, limit=10000)
    wb = Workbook()
    ws = wb.active
    ws.title = "員工清單"
    headers = [
        "員工編號(id)", "姓名", "出生年月日", "身分證字號", "戶籍地址", "居住地址", "同戶籍",
        "薪資類型", "薪資數值", "投保薪資級距", "加保日期", "退保日期", "眷屬數量", "備註",
    ]
    ws.append(headers)
    for e in employees:
        ws.append([
            e.id, e.name, e.birth_date.isoformat() if e.birth_date else "", decrypt(e.national_id) or "",
            decrypt(e.reg_address) or "", decrypt(e.live_address) or "", "是" if e.live_same_as_reg else "否",
            e.salary_type or "", float(e.salary_value) if e.salary_value else "", float(e.insured_salary_level) if e.insured_salary_level else "",
            e.enroll_date.isoformat() if e.enroll_date else "", e.cancel_date.isoformat() if e.cancel_date else "",
            e.dependent_count, (e.notes or "")[:500],
        ])
    _style_header(ws)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"employees_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/export/dependents")
async def export_dependents_excel(db: AsyncSession = Depends(get_db)):
    """匯出眷屬清單 Excel（含員工編號、姓名）"""
    employees = await crud.list_employees(db, skip=0, limit=10000, load_dependents=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "眷屬清單"
    headers = ["員工編號", "員工姓名", "眷屬姓名", "出生年月日", "身分證字號", "關係", "居住縣市", "是否身障", "身障等級", "備註"]
    ws.append(headers)
    for e in employees:
        for d in e.dependents:
            ws.append([
                e.id, e.name, d.name, d.birth_date.isoformat() if d.birth_date else "", decrypt(d.national_id) or "",
                d.relation, d.city or "", "是" if d.is_disabled else "否", d.disability_level or "", (d.notes or "")[:200],
            ])
    _style_header(ws)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"dependents_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/export/monthly-burden")
async def export_monthly_burden_excel(
    year: int = Query(..., description="年度"),
    month: int = Query(..., ge=1, le=12, description="月份"),
    db: AsyncSession = Depends(get_db),
):
    """匯出當月公司負擔明細 Excel（勞健保/職災/勞退 雇主負擔）"""
    employees = await crud.list_employees(db, skip=0, limit=10000, load_dependents=True)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月公司負擔"
    headers = [
        "員工編號", "姓名", "投保薪資級距", "眷屬人數",
        "勞保(雇主)", "健保(雇主)", "職災(雇主)", "勞退6%(雇主)", "團保",
        "公司負擔小計",
    ]
    ws.append(headers)
    rules = await crud.get_all_insurance_rules(db, year=year, month=month)
    total_employer = 0
    default_level = Decimal("26400")
    for e in employees:
        dep_count = e.dependent_count if e.dependent_count is not None else len(e.dependents or [])
        level = e.insured_salary_level or default_level
        if level <= 0:
            level = default_level
        persons = None
        if e.dependents is not None and len(e.dependents) >= 0:
            persons = [{"name": e.name, "is_employee": True, "birth_date": e.birth_date.isoformat() if e.birth_date else None, "city": None, "disability_level": None}]
            for d in e.dependents or []:
                persons.append({"name": d.name, "is_employee": False, "birth_date": d.birth_date.isoformat() if d.birth_date else None, "city": d.city, "disability_level": d.disability_level if d.is_disabled else None})
        est = estimate_insurance(
            dependent_count=dep_count,
            rules=rules,
            insured_salary_level=level,
            persons=persons,
            year=year,
            month=month,
            enroll_date=e.enroll_date,
            cancel_date=e.cancel_date,
        )
        row_total = float(est.total_employer)
        total_employer += row_total
        ws.append([
            e.id, e.name, float(level), est.dependent_count,
            float(est.labor_insurance.employer), float(est.health_insurance.employer),
            float(est.occupational_accident.employer), float(est.labor_pension.employer),
            float(est.group_insurance.employer), row_total,
        ])
    ws.append([])
    ws.append(["合計", "", "", "", "", "", "", "", "", total_employer])
    _style_header(ws)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"monthly_burden_{year}{month:02d}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/export/personal-burden")
async def export_personal_burden_excel(
    year: int = Query(..., description="年度"),
    month: int = Query(..., ge=1, le=12, description="月份"),
    db: AsyncSession = Depends(get_db),
):
    """匯出當月員工個人負擔明細 Excel（勞保+健保 個人負擔）"""
    employees = await crud.list_employees(db, skip=0, limit=10000, load_dependents=True)
    rules = await crud.get_all_insurance_rules(db, year=year, month=month)
    default_level = Decimal("26400")
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月個人負擔"
    headers = ["員工編號", "姓名", "投保薪資級距", "眷屬人數", "勞保(個人)", "健保(個人)", "個人負擔小計"]
    ws.append(headers)
    for e in employees:
        dep_count = e.dependent_count if e.dependent_count is not None else len(e.dependents or [])
        level = e.insured_salary_level or default_level
        if level <= 0:
            level = default_level
        persons = None
        if e.dependents is not None:
            persons = [{"name": e.name, "is_employee": True, "birth_date": e.birth_date.isoformat() if e.birth_date else None, "city": None, "disability_level": None}]
            for d in e.dependents or []:
                persons.append({"name": d.name, "is_employee": False, "birth_date": d.birth_date.isoformat() if d.birth_date else None, "city": d.city, "disability_level": d.disability_level if d.is_disabled else None})
        est = estimate_insurance(
            dependent_count=dep_count,
            rules=rules,
            insured_salary_level=level,
            persons=persons,
            year=year,
            month=month,
            enroll_date=e.enroll_date,
            cancel_date=e.cancel_date,
        )
        lab_emp = float(est.labor_insurance.employee)
        health_emp = float(est.health_insurance.employee)
        ws.append([e.id, e.name, float(level), est.dependent_count, lab_emp, health_emp, lab_emp + health_emp])
    _style_header(ws)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"personal_burden_{year}{month:02d}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={filename}"})
