"""人事資料備份與還原（災難復原）。僅管理員可用。"""
from datetime import datetime, date
from decimal import Decimal
from typing import Optional
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Header
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook
from app.config import settings
from app.database import get_db
from app import crud, schemas
from app.sensitive import employee_to_read_dict, dependent_to_read_dict
from app.services.backup_job import build_hr_backup_buffer, list_backup_files, get_backup_path

router = APIRouter(prefix="/api/backup", tags=["backup-restore"])


def require_admin_token(x_admin_token: Optional[str] = None) -> None:
    """僅管理員可呼叫：須設定 admin_backup_token 且請求帶相同 X-Admin-Token。"""
    if not settings.admin_backup_token:
        raise HTTPException(
            status_code=503,
            detail="未設定管理員憑證，無法使用備份/還原。請聯絡系統管理員設定 ADMIN_BACKUP_TOKEN。",
        )
    if not x_admin_token or x_admin_token.strip() != settings.admin_backup_token.strip():
        raise HTTPException(status_code=403, detail="僅管理員可使用備份與還原功能。")


# Excel 欄位順序（與 model 對應，還原時依此讀取）
EMPLOYEE_COLUMNS = [
    "id", "name", "birth_date", "national_id", "reg_address", "live_address",
    "live_same_as_reg", "salary_type", "salary_value", "insured_salary_level",
    "enroll_date", "cancel_date", "dependent_count", "safety_pdf_path",
    "contract_84_1_pdf_path", "notes", "created_at", "updated_at",
]
DEPENDENT_COLUMNS = [
    "id", "employee_id", "name", "birth_date", "national_id", "relation",
    "city", "is_disabled", "disability_level", "notes", "created_at", "updated_at",
]


def _cell_value(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.isoformat() if hasattr(v, "isoformat") else str(v)
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, Decimal):
        return str(v)
    return str(v)


def _parse_cell(sheet, row: int, col: int, col_name: str):
    c = sheet.cell(row=row, column=col)
    val = c.value
    if val is None or val == "":
        return None
    if col_name in ("birth_date", "enroll_date", "cancel_date", "created_at", "updated_at"):
        if isinstance(val, datetime):
            return val.date() if hasattr(val, "date") else val
        if isinstance(val, date) and not isinstance(val, datetime):
            return val
        s = str(val).strip()
        if not s:
            return None
        s = s[:10]
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None
    if col_name in ("live_same_as_reg", "is_disabled"):
        return str(val).strip() in ("1", "true", "True", "yes", "Y")
    if col_name in ("salary_value", "insured_salary_level"):
        try:
            return Decimal(str(val))
        except Exception:
            return None
    if col_name in ("dependent_count", "id", "employee_id"):
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return None
    return str(val).strip() if val else None


@router.get("/export")
async def export_hr_backup(
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token"),
    db: AsyncSession = Depends(get_db),
):
    """匯出完整人事資料為 Excel（員工 + 眷屬，每表一 Sheet）。檔名 hr_backup_YYYYMMDD_HHMMSS.xlsx。"""
    require_admin_token(x_admin_token)
    employees = await crud.list_employees(
        db, skip=0, limit=100000, load_dependents=True, search=None
    )
    buf, filename = build_hr_backup_buffer(employees)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/history")
async def backup_history(
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token"),
):
    """列出歷史備份檔（server/backup/hr 內 hr_backup_*.xlsx），依時間由新到舊。僅管理員。"""
    require_admin_token(x_admin_token)
    return list_backup_files()


@router.get("/download/{filename}")
async def download_backup(
    filename: str,
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token"),
):
    """下載指定歷史備份檔。僅管理員。檔名須符合白名單 hr_backup_YYYYMMDD_HHMMSS.xlsx，防路徑穿越。"""
    require_admin_token(x_admin_token)
    path = get_backup_path(filename)
    if not path:
        raise HTTPException(status_code=404, detail="備份檔不存在或檔名不正確。")
    return FileResponse(
        path,
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.post("/restore")
async def restore_hr_backup(
    file: UploadFile = File(...),
    confirm: str = Form(..., description="輸入 yes 確認覆蓋現有資料"),
    x_admin_token: Optional[str] = Header(None, alias="X-Admin-Token"),
    db: AsyncSession = Depends(get_db),
):
    """上傳本系統匯出的 Excel，清空現有員工與眷屬後寫入。需二次確認（confirm=yes）。"""
    require_admin_token(x_admin_token)
    if confirm.strip().lower() != "yes":
        raise HTTPException(status_code=400, detail="還原需二次確認，請在確認欄輸入 yes。")

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="請上傳 .xlsx 檔案。")

    try:
        content = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"讀取檔案失敗：{e}")

    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if "employees" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Excel 須包含 employees 工作表。")
    ws_emp = wb["employees"]
    ws_dep = wb["dependents"] if "dependents" in wb.sheetnames else None

    # 讀取員工列（第一列為標題）
    emp_rows = []
    for r in range(2, ws_emp.max_row + 1):
        row = {}
        for col, name in enumerate(EMPLOYEE_COLUMNS, 1):
            row[name] = _parse_cell(ws_emp, r, col, name)
        if row.get("name"):
            emp_rows.append(row)

    # 讀取眷屬列
    dep_rows = []
    if ws_dep:
        for r in range(2, ws_dep.max_row + 1):
            row = {}
            for col, name in enumerate(DEPENDENT_COLUMNS, 1):
                row[name] = _parse_cell(ws_dep, r, col, name)
            if row.get("employee_id") is not None and row.get("name"):
                dep_rows.append(row)

    # 清空現有員工（cascade 會刪除眷屬等）
    await crud.delete_all_employees(db)
    await db.commit()

    # 建立員工並建立 old_id -> new_id 對照
    old_to_new: dict[int, int] = {}
    for erow in emp_rows:
        old_id = erow.get("id")
        data = schemas.EmployeeCreate(
            name=erow.get("name") or "",
            birth_date=erow.get("birth_date") or date(1990, 1, 1),
            national_id=erow.get("national_id") or "",
            reg_address=erow.get("reg_address") or "",
            live_address=erow.get("live_address") or "",
            live_same_as_reg=bool(erow.get("live_same_as_reg")),
            salary_type=erow.get("salary_type"),
            salary_value=Decimal(str(erow["salary_value"])) if erow.get("salary_value") is not None else None,
            insured_salary_level=Decimal(str(erow["insured_salary_level"])) if erow.get("insured_salary_level") is not None else None,
            enroll_date=erow.get("enroll_date"),
            cancel_date=erow.get("cancel_date"),
            dependent_count=int(erow["dependent_count"]) if erow.get("dependent_count") is not None else 0,
            notes=erow.get("notes"),
            dependents=None,
        )
        emp = await crud.create_employee(db, data)
        await db.flush()
        if old_id is not None:
            old_to_new[int(old_id)] = emp.id

    # 建立眷屬（依 employee_id 對應到新 id）
    dep_count = 0
    for drow in dep_rows:
        old_emp_id = drow.get("employee_id")
        new_emp_id = old_to_new.get(old_emp_id) if old_emp_id is not None else None
        if new_emp_id is None:
            continue
        dep_data = schemas.DependentCreate(
            name=drow.get("name") or "",
            birth_date=drow.get("birth_date"),
            national_id=drow.get("national_id"),
            relation=drow.get("relation") or "其他",
            city=drow.get("city"),
            is_disabled=bool(drow.get("is_disabled")),
            disability_level=drow.get("disability_level"),
            notes=drow.get("notes"),
        )
        await crud.create_dependent(db, new_emp_id, dep_data)
        dep_count += 1

    await db.commit()
    return {
        "message": "還原完成",
        "restored_employees": len(emp_rows),
        "restored_dependents": dep_count,
    }
