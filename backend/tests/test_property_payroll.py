from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from starlette.datastructures import UploadFile
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool

from app.accounting.holiday_calendar import get_holiday_dates
from app.accounting.security_payroll_service import SecurityPayrollCalculator, get_holiday_count
from app.database import Base
from app.models import AccountingPayrollResult, Employee, Site
from app.routers import accounting as accounting_router


def _error_messages(errors):
    msgs = []
    for e in errors:
        if isinstance(e, dict):
            msgs.append(str(e.get("message", "")))
        else:
            msgs.append(str(e))
    return msgs


@pytest.fixture
async def async_session():
    engine = create_async_engine(
        "sqlite+aiosqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    try:
        yield session_factory
    finally:
        await engine.dispose()


def _month_weeks(year: int, month: int) -> list[date]:
    first_day = date(year, month, 1)
    weeks = []
    seen = set()
    d = first_day
    while d.month == month:
        key = (d.isocalendar().year, d.isocalendar().week)
        if key not in seen:
            seen.add(key)
            weeks.append(d)
        d += timedelta(days=1)
    return weeks


@pytest.mark.asyncio
async def test_property_weekly_and_missing_salary(async_session):
    async with async_session() as db:
        db.add_all(
            [
                Site(
                    name="物業A",
                    client_name="客戶A",
                    address="地址A",
                    contract_start=date(2025, 1, 1),
                    monthly_amount=Decimal("100000"),
                    payment_method="transfer",
                    receivable_day=10,
                    is_84_1=False,
                ),
                Site(
                    name="物業B",
                    client_name="客戶B",
                    address="地址B",
                    contract_start=date(2025, 1, 1),
                    monthly_amount=Decimal("100000"),
                    payment_method="transfer",
                    receivable_day=10,
                    is_84_1=False,
                ),
                Site(
                    name="物業C",
                    client_name="客戶C",
                    address="地址C",
                    contract_start=date(2025, 1, 1),
                    monthly_amount=Decimal("100000"),
                    payment_method="transfer",
                    receivable_day=10,
                    is_84_1=False,
                ),
                Employee(
                    name="物業達標",
                    birth_date=date(1990, 1, 1),
                    national_id="B123456789",
                    reg_address="台北",
                    live_address="台北",
                    live_same_as_reg=True,
                    registration_type="property",
                    property_pay_mode="WEEKLY_2H",
                    property_salary=Decimal("50000"),
                    weekly_amount=Decimal("10000"),
                ),
                Employee(
                    name="物業缺週",
                    birth_date=date(1991, 1, 1),
                    national_id="B223456789",
                    reg_address="台中",
                    live_address="台中",
                    live_same_as_reg=True,
                    registration_type="property",
                    property_pay_mode="WEEKLY_2H",
                    property_salary=Decimal("50000"),
                    weekly_amount=Decimal("10000"),
                ),
                Employee(
                    name="物業未設薪資",
                    birth_date=date(1992, 1, 1),
                    national_id="B323456789",
                    reg_address="桃園",
                    live_address="桃園",
                    live_same_as_reg=True,
                    registration_type="property",
                    property_pay_mode="WEEKLY_2H",
                    property_salary=None,
                    weekly_amount=None,
                ),
                Employee(
                    name="物業未設模式",
                    birth_date=date(1993, 1, 1),
                    national_id="B423450001",
                    reg_address="新北",
                    live_address="新北",
                    live_same_as_reg=True,
                    registration_type="property",
                    property_pay_mode=None,
                    property_salary=Decimal("26000"),
                    weekly_amount=Decimal("26000"),
                ),
            ]
        )
        await db.commit()

    rows = []
    week_dates = [date(2026, 1, 1), date(2026, 1, 5), date(2026, 1, 12), date(2026, 1, 19), date(2026, 1, 26)]
    for d in week_dates:
        rows.append({"site": "物業A", "employee": "物業達標", "date": datetime(d.year, d.month, d.day), "hours": 2})
    for d in week_dates[:2]:
        rows.append({"site": "物業B", "employee": "物業缺週", "date": datetime(d.year, d.month, d.day), "hours": 2})
    for d in week_dates:
        rows.append({"site": "物業C", "employee": "物業未設薪資", "date": datetime(d.year, d.month, d.day), "hours": 2})
    for d in week_dates:
        rows.append({"site": "物業A", "employee": "物業未設模式", "date": datetime(d.year, d.month, d.day), "hours": 2})

    async with async_session() as db:
        calculator = SecurityPayrollCalculator(db)
        results, errors, _ = await calculator.validate_and_calculate(rows, year=2026, month=1, payroll_type="property")
        by_name = {r["employee"]: r for r in results}

        assert by_name["物業達標"]["gross_salary"] == 50000
        assert by_name["物業達標"]["status"] == "每週2小時：完成 5/5 週"
        assert by_name["物業缺週"]["gross_salary"] == 20000
        assert by_name["物業缺週"]["deductions_total"] == 0
        assert by_name["物業缺週"]["status"] == "每週2小時：完成 2/5 週"
        assert by_name["物業未設薪資"]["gross_salary"] == 0
        assert by_name["物業未設薪資"]["deductions_total"] == 0
        assert by_name["物業未設薪資"]["net_salary"] == 0
        assert by_name["物業未設薪資"]["status"] == "未設定物業薪資"
        assert "物業未設模式" not in by_name
        missing_cfg = next((e for e in errors if isinstance(e, dict) and e.get("type") == "missing_pay_config"), None)
        assert missing_cfg is not None
        assert missing_cfg.get("employee_name") == "物業未設模式"
        assert missing_cfg.get("employee_id") is not None
        assert missing_cfg.get("current_payroll_type") == "property"
        msgs = _error_messages(errors)
        assert "員工【物業未設模式】未設定物業計薪模式" in msgs
        assert all("未設定物業薪資" not in e for e in msgs)


@pytest.mark.asyncio
async def test_property_monthly_8h_holiday(async_session):
    async with async_session() as db:
        db.add_all(
            [
                Site(
                    name="物業M",
                    client_name="客戶M",
                    address="地址M",
                    contract_start=date(2025, 1, 1),
                    monthly_amount=Decimal("100000"),
                    payment_method="transfer",
                    receivable_day=10,
                    is_84_1=False,
                ),
                Site(
                    name="物業N",
                    client_name="客戶N",
                    address="地址N",
                    contract_start=date(2025, 1, 1),
                    monthly_amount=Decimal("100000"),
                    payment_method="transfer",
                    receivable_day=10,
                    is_84_1=False,
                ),
                Employee(
                    name="物業月達標",
                    birth_date=date(1990, 2, 1),
                    national_id="B423456789",
                    reg_address="台北",
                    live_address="台北",
                    live_same_as_reg=True,
                    registration_type="property",
                    property_pay_mode="MONTHLY_8H_HOLIDAY",
                    property_salary=Decimal("36000"),
                ),
                Employee(
                    name="物業月不足",
                    birth_date=date(1991, 2, 1),
                    national_id="B523456789",
                    reg_address="高雄",
                    live_address="高雄",
                    live_same_as_reg=True,
                    registration_type="property",
                    property_pay_mode="MONTHLY_8H_HOLIDAY",
                    property_salary=Decimal("36000"),
                ),
            ]
        )
        await db.commit()

    required_days = 28 - get_holiday_count(2026, 2)
    assert required_days > 0

    holiday_dates = get_holiday_dates(2026, 2)
    working_dates = [date(2026, 2, d) for d in range(1, 29) if date(2026, 2, d) not in holiday_dates]
    assert len(working_dates) == required_days

    rows = []
    for d in working_dates:
        rows.append({"site": "物業M", "employee": "物業月達標", "date": datetime(d.year, d.month, d.day), "hours": 8})
    for d in working_dates[:-1]:
        rows.append({"site": "物業N", "employee": "物業月不足", "date": datetime(d.year, d.month, d.day), "hours": 8})

    async with async_session() as db:
        calculator = SecurityPayrollCalculator(db)
        results, _, _ = await calculator.validate_and_calculate(rows, year=2026, month=2, payroll_type="property")
        by_name = {r["employee"]: r for r in results}

        assert by_name["物業月達標"]["gross_salary"] == 36000
        assert by_name["物業月達標"]["status"] == f"滿班（{required_days}天）"
        expected_partial = int(round(36000 * ((required_days - 1) / required_days), 0))
        assert by_name["物業月不足"]["gross_salary"] == expected_partial
        assert by_name["物業月不足"]["deductions_total"] == 0
        assert by_name["物業月不足"]["status"] == f"出勤 {required_days - 1}/{required_days}（按比例）"


@pytest.mark.asyncio
async def test_property_history_and_export(async_session):
    async with async_session() as db:
        db.add(
            AccountingPayrollResult(
                year=2026,
                month=1,
                type="property",
                site="物業站點",
                employee="物業人員",
                pay_type="monthly",
                total_hours=160,
                gross_salary=32000,
                labor_insurance_employee=0,
                health_insurance_employee=0,
                group_insurance=0,
                self_pension_6=0,
                deductions_total=0,
                net_salary=32000,
                total_salary=32000,
                status="達標（8小時/日）",
            )
        )
        await db.commit()

    async with async_session() as db:
        history = await accounting_router.security_payroll_history(
            year=2026, month=1, payroll_type="property", db=db
        )
        assert history["payroll_type"] == "property"
        assert history["summary"]["row_count"] == 1

        response = await accounting_router.security_payroll_export_get(
            year=2026, month=1, payroll_type="property", db=db
        )
        content = b""
        async for chunk in response.body_iterator:
            content += chunk
        wb = load_workbook(BytesIO(content))
        assert "全部顯示" in wb.sheetnames


@pytest.mark.asyncio
async def test_property_weekly_amount_five_weeks(async_session):
    async with async_session() as db:
        db.add(
            Site(
                name="驗收案場",
                client_name="驗收客戶",
                address="驗收地址",
                contract_start=date(2025, 1, 1),
                monthly_amount=Decimal("100000"),
                payment_method="transfer",
                receivable_day=10,
                is_84_1=False,
            )
        )
        db.add(
            Employee(
                name="驗收物業",
                birth_date=date(1990, 1, 1),
                national_id="Z223456789",
                reg_address="台北",
                live_address="台北",
                live_same_as_reg=True,
                registration_type="property",
                property_pay_mode="WEEKLY_2H",
                property_salary=Decimal("50000"),
                weekly_amount=Decimal("10000"),
            )
        )
        await db.commit()

    rows = []
    # 2026-01 橫跨 5 個 ISO 週：1/1、1/5、1/12、1/19、1/26
    for d in [date(2026, 1, 1), date(2026, 1, 5), date(2026, 1, 12), date(2026, 1, 19), date(2026, 1, 26)]:
        rows.append({"site": "驗收案場", "employee": "驗收物業", "date": datetime(d.year, d.month, d.day), "hours": 2})

    async with async_session() as db:
        calculator = SecurityPayrollCalculator(db)
        results, _, _ = await calculator.validate_and_calculate(rows, year=2026, month=1, payroll_type="property")
        assert len(results) == 1
        row = results[0]
        assert row["gross_salary"] == 50000
        assert row["status"] == "每週2小時：完成 5/5 週"


@pytest.mark.asyncio
async def test_property_weekly_cross_site_hours_merged(async_session):
    async with async_session() as db:
        db.add_all(
            [
                Site(
                    name="跨站A",
                    client_name="客戶A",
                    address="地址A",
                    contract_start=date(2025, 1, 1),
                    monthly_amount=Decimal("100000"),
                    payment_method="transfer",
                    receivable_day=10,
                    is_84_1=False,
                ),
                Site(
                    name="跨站B",
                    client_name="客戶B",
                    address="地址B",
                    contract_start=date(2025, 1, 1),
                    monthly_amount=Decimal("100000"),
                    payment_method="transfer",
                    receivable_day=10,
                    is_84_1=False,
                ),
                Employee(
                    name="跨站物業",
                    birth_date=date(1990, 1, 1),
                    national_id="Z323456789",
                    reg_address="台北",
                    live_address="台北",
                    live_same_as_reg=True,
                    registration_type="property",
                    property_pay_mode="WEEKLY_2H",
                    property_salary=Decimal("50000"),
                    weekly_amount=Decimal("10000"),
                ),
            ]
        )
        await db.commit()

    rows = [
        {"site": "跨站A", "employee": "跨站物業", "date": datetime(2026, 1, 5), "hours": 1},
        {"site": "跨站B", "employee": "跨站物業", "date": datetime(2026, 1, 6), "hours": 1},
        {"site": "跨站A", "employee": "跨站物業", "date": datetime(2026, 1, 12), "hours": 2},
        {"site": "跨站A", "employee": "跨站物業", "date": datetime(2026, 1, 19), "hours": 2},
        {"site": "跨站A", "employee": "跨站物業", "date": datetime(2026, 1, 26), "hours": 2},
        {"site": "跨站A", "employee": "跨站物業", "date": datetime(2026, 1, 1), "hours": 2},
    ]

    async with async_session() as db:
        calculator = SecurityPayrollCalculator(db)
        results, _, _ = await calculator.validate_and_calculate(rows, year=2026, month=1, payroll_type="property")
        rows_by_employee = [r for r in results if r["employee"] == "跨站物業"]
        assert len(rows_by_employee) == 2
        gross_values = sorted([r["gross_salary"] for r in rows_by_employee], reverse=True)
        assert gross_values == [25000, 25000]
        assert sum(r["gross_salary"] for r in rows_by_employee) == 50000


@pytest.mark.asyncio
async def test_property_weekly_split_three_sites(async_session):
    async with async_session() as db:
        db.add_all(
            [
                Site(
                    name="三本公園",
                    client_name="客戶1",
                    address="地址1",
                    contract_start=date(2025, 1, 1),
                    monthly_amount=Decimal("100000"),
                    payment_method="transfer",
                    receivable_day=10,
                    is_84_1=False,
                ),
                Site(
                    name="西雅圖",
                    client_name="客戶2",
                    address="地址2",
                    contract_start=date(2025, 1, 1),
                    monthly_amount=Decimal("100000"),
                    payment_method="transfer",
                    receivable_day=10,
                    is_84_1=False,
                ),
                Site(
                    name="長榮富豪",
                    client_name="客戶3",
                    address="地址3",
                    contract_start=date(2025, 1, 1),
                    monthly_amount=Decimal("100000"),
                    payment_method="transfer",
                    receivable_day=10,
                    is_84_1=False,
                ),
                Employee(
                    name="張興鐘",
                    birth_date=date(1990, 1, 1),
                    national_id="Q123456789",
                    reg_address="台北",
                    live_address="台北",
                    live_same_as_reg=True,
                    registration_type="property",
                    property_pay_mode="WEEKLY_2H",
                    property_salary=Decimal("50000"),
                    weekly_amount=Decimal("10000"),
                ),
            ]
        )
        await db.commit()

    rows = [
        {"site": "三本公園", "employee": "張興鐘", "date": datetime(2026, 1, 1), "hours": 2},
        {"site": "三本公園", "employee": "張興鐘", "date": datetime(2026, 1, 5), "hours": 2},
        {"site": "西雅圖", "employee": "張興鐘", "date": datetime(2026, 1, 12), "hours": 2},
        {"site": "西雅圖", "employee": "張興鐘", "date": datetime(2026, 1, 19), "hours": 2},
        {"site": "長榮富豪", "employee": "張興鐘", "date": datetime(2026, 1, 26), "hours": 2},
    ]

    async with async_session() as db:
        calculator = SecurityPayrollCalculator(db)
        results, _, _ = await calculator.validate_and_calculate(rows, year=2026, month=1, payroll_type="property")
        by_site = {r["site"]: r for r in results if r["employee"] == "張興鐘"}
        assert by_site["三本公園"]["gross_salary"] == 16667
        assert by_site["西雅圖"]["gross_salary"] == 16667
        assert by_site["長榮富豪"]["gross_salary"] == 16666
        assert sum(v["gross_salary"] for v in by_site.values()) == 50000
        assert by_site["三本公園"]["status"] == "每週2小時：完成 5/5 週 跨案場計算（保險僅扣一次）"


@pytest.mark.asyncio
async def test_property_cross_type_lookup_with_extra(async_session):
    async with async_session() as db:
        db.add(
            Site(
                name="跨公司案場",
                client_name="客戶X",
                address="地址X",
                contract_start=date(2025, 1, 1),
                monthly_amount=Decimal("100000"),
                payment_method="transfer",
                receivable_day=10,
                is_84_1=False,
            )
        )
        db.add(
            Employee(
                name="跨公司員工",
                birth_date=date(1990, 1, 1),
                national_id="M123456789",
                reg_address="台北",
                live_address="台北",
                live_same_as_reg=True,
                registration_type="smith",
                property_pay_mode="WEEKLY_2H",
                property_salary=Decimal("50000"),
                weekly_amount=Decimal("10000"),
            )
        )
        await db.commit()

    rows = [{"site": "跨公司案場", "employee": "跨公司員工", "date": datetime(2026, 1, 5), "hours": 2}]
    async with async_session() as db:
        calculator = SecurityPayrollCalculator(db)
        results, errors, _ = await calculator.validate_and_calculate(
            rows, year=2026, month=1, payroll_type="property", extra_payroll_types=["smith"]
        )
        assert len(results) == 1
        assert results[0]["employee"] == "跨公司員工"
        assert "來源：史密斯" in results[0]["status"]
        assert all("未建立" not in e for e in _error_messages(errors))


@pytest.mark.asyncio
async def test_property_cross_type_lookup_without_extra_should_error(async_session):
    async with async_session() as db:
        db.add(
            Site(
                name="跨公司案場B",
                client_name="客戶Y",
                address="地址Y",
                contract_start=date(2025, 1, 1),
                monthly_amount=Decimal("100000"),
                payment_method="transfer",
                receivable_day=10,
                is_84_1=False,
            )
        )
        db.add(
            Employee(
                name="跨公司員工B",
                birth_date=date(1990, 1, 1),
                national_id="M223456789",
                reg_address="台北",
                live_address="台北",
                live_same_as_reg=True,
                registration_type="smith",
                property_pay_mode="WEEKLY_2H",
                property_salary=Decimal("50000"),
                weekly_amount=Decimal("10000"),
            )
        )
        await db.commit()

    rows = [{"site": "跨公司案場B", "employee": "跨公司員工B", "date": datetime(2026, 1, 5), "hours": 2}]
    async with async_session() as db:
        calculator = SecurityPayrollCalculator(db)
        results, errors, _ = await calculator.validate_and_calculate(
            rows, year=2026, month=1, payroll_type="property", extra_payroll_types=[]
        )
        assert results == []
        assert "員工【跨公司員工B】未建立" in _error_messages(errors)


@pytest.mark.asyncio
async def test_upload_endpoint_cross_type_lookup_case_a(async_session, monkeypatch):
    async with async_session() as db:
        db.add(
            Site(
                name="上傳跨公司案場",
                client_name="客戶Z",
                address="地址Z",
                contract_start=date(2025, 1, 1),
                monthly_amount=Decimal("100000"),
                payment_method="transfer",
                receivable_day=10,
                is_84_1=False,
            )
        )
        db.add(
            Employee(
                name="上傳跨公司員工",
                birth_date=date(1990, 1, 1),
                national_id="U123456789",
                reg_address="台北",
                live_address="台北",
                live_same_as_reg=True,
                registration_type="smith",
                property_pay_mode="WEEKLY_2H",
                property_salary=Decimal("50000"),
                weekly_amount=Decimal("10000"),
            )
        )
        await db.commit()

        def _fake_parse(_: bytes, __: str, year: int | None = None, month: int | None = None):
            return [
                {
                    "site": "上傳跨公司案場",
                    "employee": "上傳跨公司員工",
                    "date": datetime(2026, 1, 5),
                    "hours": 2,
                }
            ], []

        monkeypatch.setattr(accounting_router, "parse_security_hours_file", _fake_parse)
        upload = UploadFile(filename="hours.xlsx", file=BytesIO(b"dummy"))
        resp = await accounting_router.security_payroll_upload(
            file=upload,
            year=2026,
            month=1,
            type="property",
            payroll_type="property",
            extra_payroll_types='["smith"]',
            db=db,
        )
        assert len(resp["results"]) == 1
        assert all("未建立" not in e for e in _error_messages(resp["errors"]))


@pytest.mark.asyncio
async def test_upload_endpoint_cross_type_lookup_case_b(async_session, monkeypatch):
    async with async_session() as db:
        db.add(
            Site(
                name="上傳跨公司案場B",
                client_name="客戶W",
                address="地址W",
                contract_start=date(2025, 1, 1),
                monthly_amount=Decimal("100000"),
                payment_method="transfer",
                receivable_day=10,
                is_84_1=False,
            )
        )
        db.add(
            Employee(
                name="上傳跨公司員工B",
                birth_date=date(1990, 1, 1),
                national_id="U223456789",
                reg_address="台北",
                live_address="台北",
                live_same_as_reg=True,
                registration_type="smith",
                property_pay_mode="WEEKLY_2H",
                property_salary=Decimal("50000"),
                weekly_amount=Decimal("10000"),
            )
        )
        await db.commit()

        def _fake_parse(_: bytes, __: str, year: int | None = None, month: int | None = None):
            return [
                {
                    "site": "上傳跨公司案場B",
                    "employee": "上傳跨公司員工B",
                    "date": datetime(2026, 1, 5),
                    "hours": 2,
                }
            ], []

        monkeypatch.setattr(accounting_router, "parse_security_hours_file", _fake_parse)
        upload = UploadFile(filename="hours.xlsx", file=BytesIO(b"dummy"))
        resp = await accounting_router.security_payroll_upload(
            file=upload,
            year=2026,
            month=1,
            type="property",
            payroll_type="property",
            extra_payroll_types="[]",
            db=db,
        )
        assert resp["results"] == []
        assert "員工【上傳跨公司員工B】未建立" in _error_messages(resp["errors"])


@pytest.mark.asyncio
async def test_cross_company_property_monthly_mode_should_calculate(async_session):
    async with async_session() as db:
        db.add(
            Site(
                name="跨公司物業月薪案場",
                client_name="客戶M",
                address="地址M",
                contract_start=date(2025, 1, 1),
                monthly_amount=Decimal("100000"),
                payment_method="transfer",
                receivable_day=10,
                is_84_1=False,
            )
        )
        db.add(
            Employee(
                name="林憶慧",
                birth_date=date(1990, 1, 1),
                national_id="N123456789",
                reg_address="台北",
                live_address="台北",
                live_same_as_reg=True,
                registration_type="smith",
                property_pay_mode="monthly",
                salary_value=Decimal("36000"),
                insured_salary_level=Decimal("36000"),
            )
        )
        await db.commit()

    rows = [{"site": "跨公司物業月薪案場", "employee": "林憶慧", "date": datetime(2026, 1, 5), "hours": 12}]
    async with async_session() as db:
        calculator = SecurityPayrollCalculator(db)
        results, errors, _ = await calculator.validate_and_calculate(
            rows,
            year=2026,
            month=1,
            payroll_type="property",
            extra_payroll_types=["smith"],
        )
        assert len(results) == 1
        assert results[0]["employee"] == "林憶慧"
        assert results[0]["gross_salary"] > 0
        assert all("未設定物業計薪模式" not in m for m in _error_messages(errors))
