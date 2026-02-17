from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool

from app.accounting.payroll_export import build_payroll_excel, build_payroll_excel_grouped
from app.database import Base
from app.models import AccountingPayrollResult, Employee, Site, SiteEmployeeAssignment
from app.routers import accounting as accounting_router


@pytest.fixture
async def async_session():
    engine = create_async_engine(
        "sqlite+aiosqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
    try:
        yield session_factory
    finally:
        await engine.dispose()


async def _seed_history_row(db: AsyncSession, site: str, employee: str):
    db.add(
        AccountingPayrollResult(
            year=2026,
            month=1,
            type="security",
            site=site,
            employee=employee,
            pay_type="monthly",
            total_hours=240,
            gross_salary=42000,
            labor_insurance_employee=900,
            health_insurance_employee=650,
            group_insurance=350,
            self_pension_6=0,
            deductions_total=1900,
            net_salary=40100,
            total_salary=40100,
            status="滿班",
        )
    )


@pytest.mark.asyncio
async def test_history_api_enrich_and_stats(async_session):
    async with async_session() as db:
        site = Site(
            name="測試案場A",
            client_name="客戶A",
            address="地址A",
            contract_start=date(2025, 1, 1),
            contract_end=None,
            monthly_amount=Decimal("100000"),
            payment_method="transfer",
            receivable_day=10,
            is_84_1=False,
        )
        db.add(site)
        await db.flush()
        emp = Employee(
            name="王小明",
            birth_date=date(1990, 1, 1),
            national_id="A123456789",
            reg_address="台北",
            live_address="台北",
            live_same_as_reg=True,
            pay_method="SECURITY_FIRST",
            bank_code="007",
            branch_code="1234",
            bank_account="1234567890123",
        )
        db.add(emp)
        await db.flush()
        db.add(SiteEmployeeAssignment(site_id=site.id, employee_id=emp.id))
        await _seed_history_row(db, "測試案場A", "王小明")
        await db.commit()

    async with async_session() as db:
        data = await accounting_router.security_payroll_history(
            year=2026, month=1, payroll_type="security", db=db
        )
        assert data["summary"]["row_count"] == 1
        assert data["stats"]["sec_first"] == 1
        row = data["results"][0]
        assert row["salary_type"] == "保全一銀"
        assert row["bank_code"] == "007"
        assert row["branch_code"] == "1234"
        assert row["account_number"] == "1234567890123"
        assert row["conflict"] is False
        assert row["matched_candidates_count"] == 1


@pytest.mark.asyncio
async def test_history_api_unset_when_no_employee(async_session):
    async with async_session() as db:
        await _seed_history_row(db, "未知案場", "不存在員工")
        await db.commit()

    async with async_session() as db:
        data = await accounting_router.security_payroll_history(
            year=2026, month=1, payroll_type="security", db=db
        )
        row = data["results"][0]
        assert row["salary_type"] == "未設定"
        assert row["bank_code"] == ""
        assert row["branch_code"] == ""
        assert row["account_number"] == ""
        assert row["conflict"] is False
        assert row["matched_candidates_count"] == 0
        assert data["stats"]["unset"] == 1


@pytest.mark.asyncio
async def test_history_api_conflict_and_excel_columns(async_session):
    async with async_session() as db:
        site = Site(
            name="測試案場B",
            client_name="客戶B",
            address="地址B",
            contract_start=date(2025, 1, 1),
            contract_end=None,
            monthly_amount=Decimal("100000"),
            payment_method="transfer",
            receivable_day=10,
            is_84_1=False,
        )
        db.add(site)
        await db.flush()
        emp1 = Employee(
            name="同名員工",
            birth_date=date(1990, 1, 1),
            national_id="A223456789",
            reg_address="台北",
            live_address="台北",
            live_same_as_reg=True,
            pay_method="OTHER_BANK",
            bank_code="808",
            branch_code="5678",
            bank_account="999111222333",
        )
        emp2 = Employee(
            name="同名員工",
            birth_date=date(1991, 1, 1),
            national_id="A323456789",
            reg_address="台中",
            live_address="台中",
            live_same_as_reg=True,
            pay_method="CASH",
        )
        db.add_all([emp1, emp2])
        await db.flush()
        db.add_all(
            [
                SiteEmployeeAssignment(site_id=site.id, employee_id=emp1.id),
                SiteEmployeeAssignment(site_id=site.id, employee_id=emp2.id),
            ]
        )
        await _seed_history_row(db, "測試案場B", "同名員工")
        await db.commit()

    async with async_session() as db:
        data = await accounting_router.security_payroll_history(
            year=2026, month=1, payroll_type="security", db=db
        )
        row = data["results"][0]
        assert row["conflict"] is True
        assert row["matched_candidates_count"] == 2
        assert row["salary_type"] == "未設定"
        assert data["stats"]["unset"] == 1

        content = build_payroll_excel(data["results"], sheet_name="測試")
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        headers = [ws.cell(row=1, column=i).value for i in range(1, 17)]
        assert headers == [
            "案場",
            "員工",
            "薪制",
            "總工時",
            "應發",
            "勞保",
            "健保",
            "團保",
            "自提6%",
            "扣款合計",
            "實發",
            "狀態",
            "領薪方式",
            "銀行代碼",
            "分行代碼",
            "銀行帳號",
        ]


@pytest.mark.asyncio
async def test_grouped_export_sheets_match_stats(async_session):
    async with async_session() as db:
        site = Site(
            name="測試案場C",
            client_name="客戶C",
            address="地址C",
            contract_start=date(2025, 1, 1),
            contract_end=None,
            monthly_amount=Decimal("100000"),
            payment_method="transfer",
            receivable_day=10,
            is_84_1=False,
        )
        db.add(site)
        await db.flush()
        cash_emp = Employee(
            name="領現員工",
            birth_date=date(1990, 1, 1),
            national_id="A423456789",
            reg_address="高雄",
            live_address="高雄",
            live_same_as_reg=True,
            pay_method="CASH",
        )
        sec_emp = Employee(
            name="保全一銀員工",
            birth_date=date(1990, 1, 1),
            national_id="A523456789",
            reg_address="桃園",
            live_address="桃園",
            live_same_as_reg=True,
            pay_method="SECURITY_FIRST",
            bank_code="007",
            branch_code="1234",
            bank_account="111222333444",
        )
        unset_emp = Employee(
            name="未設定員工",
            birth_date=date(1990, 1, 1),
            national_id="A623456789",
            reg_address="台南",
            live_address="台南",
            live_same_as_reg=True,
            pay_method="UNKNOWN",
        )
        db.add_all([cash_emp, sec_emp, unset_emp])
        await db.flush()
        db.add_all(
            [
                SiteEmployeeAssignment(site_id=site.id, employee_id=cash_emp.id),
                SiteEmployeeAssignment(site_id=site.id, employee_id=sec_emp.id),
                SiteEmployeeAssignment(site_id=site.id, employee_id=unset_emp.id),
            ]
        )
        await _seed_history_row(db, "測試案場C", "領現員工")
        await _seed_history_row(db, "測試案場C", "保全一銀員工")
        await _seed_history_row(db, "測試案場C", "未設定員工")
        await db.commit()

    async with async_session() as db:
        data = await accounting_router.security_payroll_history(
            year=2026, month=1, payroll_type="security", db=db
        )
        content = build_payroll_excel_grouped(data["results"], data["stats"])
        wb = load_workbook(BytesIO(content))
        assert wb.sheetnames == ["全部顯示", "領現", "保全一銀", "公寓一銀", "史密斯一銀", "其他銀行", "未設定"]

        all_ws = wb["全部顯示"]
        cash_ws = wb["領現"]
        sec_ws = wb["保全一銀"]
        apt_ws = wb["公寓一銀"]
        smith_ws = wb["史密斯一銀"]
        other_ws = wb["其他銀行"]
        unset_ws = wb["未設定"]

        assert all_ws["A1"].value == "全部顯示（3人）"
        assert cash_ws["A1"].value == "領現（1人）"
        assert sec_ws["A1"].value == "保全一銀（1人）"
        assert apt_ws["A1"].value == "公寓一銀（0人）"
        assert smith_ws["A1"].value == "史密斯一銀（0人）"
        assert other_ws["A1"].value == "其他銀行（0人）"
        assert unset_ws["A1"].value == "未設定（1人）"

        assert apt_ws["A2"].value == "無資料"
        assert smith_ws["A2"].value == "無資料"
        assert other_ws["A2"].value == "無資料"

        all_sheet_count = max(all_ws.max_row - 2, 0)
        cash_sheet_count = max(cash_ws.max_row - 2, 0)
        sec_sheet_count = max(sec_ws.max_row - 2, 0)
        unset_sheet_count = max(unset_ws.max_row - 2, 0)
        assert all_sheet_count == 3
        assert cash_sheet_count + sec_sheet_count + unset_sheet_count == all_sheet_count
